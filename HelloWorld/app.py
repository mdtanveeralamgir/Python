import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#Create a workbook object
wb = xl.load_workbook('transactions.xlsx')

#access the sheet 1 from workbook
sheet = wb['Sheet1']

#access the column of a sheet using coordinate of each cell
cell = sheet['a1']
# print(cell.value) #Print the value of that cell
"""
another way to access the cell
cell = sheet.cell(1,1) #row 1 column 1
"""
#Find how many rows do we have
# print(sheet.max_row)

#retrieving data and adding data into new spreadsheet

for row in range(2, sheet.max_row + 1):
    cellValue = sheet.cell(row, 3).value
    corrected_price = cellValue * 0.9
    corrected_price_cell = sheet.cell(row, 4)

#adding a chart

#Range of values using reference class
values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4, #only selecting the 4th column
          max_col=4 #only selecting the 4th column
          )

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')
wb.save('transactions2.xlsx')


